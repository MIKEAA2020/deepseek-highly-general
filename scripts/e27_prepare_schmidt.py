#!/usr/bin/env python3
"""
E27 data preparation: extract the Schmidt et al. 2016 (Nat Biotechnol,
"the quantitative and condition-dependent E. coli proteome") supplementary
tables into analysis-ready CSVs.

Source workbook: data/schmidt2016/NIHMS65833-supplement-Supplementary_
tables.xlsx (obtained via the Europe PMC REST supplementary-files endpoint
for PMC4888949; PRIDE raw data: PXD000498).

Extracted:
  1. schmidt2016_s8_fc.csv  -- Table S8 (dataset 2, BW25113, biological
     triplicates, SafeQuant label-free quantification): per-gene protein
     fold-changes (linear medianRatio) vs. the glucose reference, all 22
     conditions, plus q-values and CVs; Uniprot accession + gene name.
  2. schmidt2016_s6_map.csv -- Table S6 (final combined table) Uniprot
     Accession -> Bnumber + Gene mapping (plus COG class), used to map the
     proteome rows onto the iJO1366 b-number panel.
  3. schmidt2016_s7_fc.csv  -- Table S7 (dataset 1, no replicates) fold
     changes vs glucose, sensitivity arm (includes anaerobic).
"""
import hashlib
from pathlib import Path

import openpyxl
import pandas as pd

SRC = Path("/home/z/my-project/data/schmidt2016/"
           "NIHMS65833-supplement-Supplementary_tables.xlsx")
OUT = Path("/home/z/my-project/data/schmidt2016")

wb = openpyxl.load_workbook(SRC, read_only=True)

# ---------------------------------------------------------------- Table S8
ws = wb["Table S8"]
rows = list(ws.iter_rows(min_row=3, values_only=True))  # row 3 = header
hdr = [str(c) if c is not None else "" for c in rows[0]]
data = [r for r in rows[1:] if r and r[0] is not None]
s8 = pd.DataFrame(data, columns=hdr)
# keep: accession, gene, medianRatio_*, qvalue_*, cv_*
keep = ([c for c in hdr if c.startswith(("Uniprot", "Gene"))]
        + [c for c in hdr if c in ("Peptides.used.for.quan",
                                   "Confidence.score")]
        + [c for c in hdr if c.startswith("medianRatio")]
        + [c for c in hdr if c.startswith("qvalue")]
        + [c for c in hdr if c.startswith("cv_")])
s8 = s8[keep]
for c in keep:
    s8[c] = pd.to_numeric(s8[c], errors="coerce") if c != "Uniprot Accession" \
        and c != "Gene" else s8[c].astype(str).str.strip()
s8 = s8.rename(columns={"Uniprot Accession": "uniprot", "Gene": "gene_name"})
print(f"[S8] rows={len(s8)}  medianRatio cols="
      f"{sum(c.startswith('medianRatio') for c in s8.columns)}  "
      f"qvalue cols={sum(c.startswith('qvalue') for c in s8.columns)}")
s8.to_csv(OUT / "schmidt2016_s8_fc.csv", index=False)

# ---------------------------------------------------------------- Table S6
ws = wb["Table S6"]
rows = list(ws.iter_rows(min_row=3, values_only=True))
hdr = [str(c) if c is not None else "" for c in rows[0]]
# final columns after the 3 replicate blocks: Gene, Bnumber, COG...
tail = [c for c in hdr if c != ""][ -6:]
print(f"[S6] tail annotation columns: {tail}")
data = [r for r in rows[1:] if r and r[0] is not None]
s6 = pd.DataFrame(data, columns=hdr)
# positional: 0 uniprot, 1 description, 2 gene (early), 3 peptides, 4 conf,
# 5 MW, 6 dataset, 7..72 the 66 condition columns, 73 gene (tail),
# 74 Bnumber, 75 COG letter, 76 COG desc, 77 COG class
pos = {name: i for i, name in enumerate(hdr) if name}
assert pos["Bnumber"] == 74 and pos["Dataset"] == 6, pos
m = s6.iloc[:, [0, 2, 74, 77]].copy()
m.columns = ["uniprot", "gene_name", "bnum", "cog_class"]
for col in ("uniprot", "gene_name", "bnum"):
    m[col] = m[col].astype(str).str.strip()
m = m[m["uniprot"].str.startswith("P") | m["uniprot"].str.startswith("Q")]
m["bnum"] = m["bnum"].where(m["bnum"].str.match(r"^[bB]\d{4}$"), "")
m = m.drop_duplicates(subset="uniprot", keep="first")
print(f"[S6] mapping rows={len(m)}  with b-numbers="
      f"{(m['bnum'] != '').sum()}")
print(m.head(8).to_string())
m.to_csv(OUT / "schmidt2016_s6_map.csv", index=False)

# spot checks against known b-numbers (UniProt -> EcoCyc/MG1655)
spot = {"P0A9G6": ("aceA", "b4015"), "P0A8V2": ("rpoB", "b3987"),
        "P0A9B2": ("gapA", "b1779"), "P0A6F5": ("groL", "b4143")}
for acc, (gene, b) in spot.items():
    row = m[m["uniprot"] == acc]
    got = (row["gene_name"].iloc[0], row["bnum"].iloc[0]) if len(row) \
        else ("MISSING", "MISSING")
    print(f"  spot {acc} {gene}->{b}: got {got} "
          f"{'OK' if got == (gene, b) else 'MISMATCH'}")

# ---------------------------------------------------------------- Table S7
ws = wb["Table S7"]
rows = list(ws.iter_rows(min_row=3, values_only=True))
hdr = [str(c) if c is not None else "" for c in rows[0]]
data = [r for r in rows[1:] if r and r[0] is not None]
s7 = pd.DataFrame(data, columns=hdr)
keep7 = ["Uniprot Accession", "Gene"] + \
    [c for c in hdr if c.startswith("medianRatio")]
s7 = s7[keep7].rename(
    columns={"Uniprot Accession": "uniprot", "Gene": "gene_name"})
s7["uniprot"] = s7["uniprot"].astype(str).str.strip()
s7["gene_name"] = s7["gene_name"].astype(str).str.strip()
for c in keep7[2:]:
    s7[c] = pd.to_numeric(s7[c], errors="coerce")
print(f"[S7] rows={len(s7)}  conditions="
      f"{sum(c.startswith('medianRatio') for c in s7.columns)}")
s7.to_csv(OUT / "schmidt2016_s7_fc.csv", index=False)

# provenance
sha = hashlib.sha256(SRC.read_bytes()).hexdigest()
print(f"\nsha256 {SRC.name}: {sha}")
print(f"sha256 s8_fc.csv: "
      f"{hashlib.sha256((OUT / 'schmidt2016_s8_fc.csv').read_bytes()).hexdigest()}")
